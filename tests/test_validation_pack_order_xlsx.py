"""Regression test for XLSX order preservation in validation pack exports."""

from pathlib import Path
import json

from openpyxl import load_workbook

from metricas_lattes.exports.validation_pack import (
    COLUMN_ORDER,
    generate_validation_pack,
)


def _extract_ids_from_sheet(sheet) -> list[tuple[str, int | None, str | None]]:
    header = [cell.value for cell in sheet[1]]
    col_index = {name: idx for idx, name in enumerate(header)}
    ids: list[tuple[str, int | None, str | None]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        section = row[col_index['section']]
        numero_item = row[col_index['numero_item']]
        titulo = row[col_index['titulo']]
        if section is None and numero_item is None and titulo is None:
            continue
        ids.append((section or '', numero_item, titulo))
    return ids


def test_validation_pack_preserves_input_order_xlsx(tmp_path: Path) -> None:
    input_dir = tmp_path / 'input'
    output_dir = tmp_path / 'out'
    input_dir.mkdir()

    data = {
        'researcher': {'lattes_id': '123', 'full_name': 'Tester'},
        'productions': [
            {
                'numero_item': 1,
                'titulo': 'A1',
                'source': {'production_type': 'Artigos'},
            },
            {
                'numero_item': 1,
                'titulo': 'B1',
                'source': {'production_type': 'Livros'},
            },
            {
                'numero_item': 2,
                'titulo': 'A2',
                'source': {'production_type': 'Artigos'},
            },
            {
                'numero_item': 2,
                'titulo': 'B2',
                'source': {'production_type': 'Livros'},
            },
        ],
        'metadata': {
            'sections': [
                {'section_title': 'Artigos'},
                {'section_title': 'Livros'},
            ]
        },
    }

    json_path = input_dir / '123__tester.json'
    json_path.write_text(json.dumps(data, ensure_ascii=False), encoding='utf-8')

    generate_validation_pack(input_dir, output_dir, ['xlsx'])

    xlsx_path = output_dir / 'researchers' / '123__' / 'VALIDACAO.xlsx'
    workbook = load_workbook(xlsx_path, data_only=True)
    assert 'Produções' in workbook.sheetnames

    sheet = workbook['Produções']
    header = [cell.value for cell in sheet[1]]
    assert header == COLUMN_ORDER

    expected = [
        (item['source']['production_type'], item['numero_item'], item['titulo'])
        for item in data['productions']
    ]
    observed = _extract_ids_from_sheet(sheet)

    assert observed == expected
