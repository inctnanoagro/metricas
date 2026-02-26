from __future__ import annotations

import html as html_lib
import json
import re
import subprocess
import sys
from pathlib import Path

import pytest

from metricas_lattes.parser_router import normalize_filename


DATA_DIR = Path("data/full_profiles_20250114")
SCHEMA_PATH = Path("schema/researcher_output.schema.json")
TARGET_SECTIONS = {
    "artigos completos publicados em periodicos",
    "capitulos de livros publicados",
    "textos em jornais de noticias revistas",
}


def _select_diverse_profiles(input_dir: Path) -> list[Path]:
    html_files = [
        path
        for path in input_dir.glob("*.html")
        if path.is_file() and not path.name.startswith("._")
    ]
    html_files.sort(key=lambda path: path.stat().st_size)
    if not html_files:
        return []
    indices = [0, len(html_files) // 2, len(html_files) - 1]
    selected = []
    seen = set()
    for index in indices:
        if 0 <= index < len(html_files):
            candidate = html_files[index]
            if candidate not in seen:
                selected.append(candidate)
                seen.add(candidate)
    return selected


def _load_schema() -> dict:
    return json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))


def _validate_schema(payload: dict, schema: dict) -> list[str]:
    from jsonschema import Draft202012Validator

    validator = Draft202012Validator(schema)
    errors = []
    for error in sorted(validator.iter_errors(payload), key=lambda err: list(err.path)):
        path = ".".join(str(part) for part in error.path)
        errors.append(f"{path}: {error.message}" if path else error.message)
    return errors


def _normalize_section(value: str) -> str:
    normalized = normalize_filename(value)
    normalized = normalized.replace("/", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


@pytest.mark.integration
def test_full_pipeline_validation_pack(tmp_path: Path) -> None:
    if not DATA_DIR.exists():
        pytest.skip(f"Data directory not found: {DATA_DIR}")

    selected = _select_diverse_profiles(DATA_DIR)
    if len(selected) < 3:
        pytest.skip("Not enough full_profile HTML files for integration test")

    temp_in = tmp_path / "input"
    temp_out = tmp_path / "out"
    temp_in.mkdir(parents=True, exist_ok=True)

    for path in selected:
        (temp_in / path.name).write_bytes(path.read_bytes())

    batch_cmd = [
        sys.executable,
        "-m",
        "metricas_lattes.batch_full_profile",
        "--in",
        str(temp_in),
        "--out",
        str(temp_out),
        "--schema",
        str(SCHEMA_PATH),
        "--years",
        "all",
    ]
    batch_run = subprocess.run(batch_cmd, capture_output=True, text=True)
    assert batch_run.returncode == 0, batch_run.stdout + batch_run.stderr

    researchers_dir = temp_out / "researchers"
    json_paths = sorted(researchers_dir.glob("*.json"))
    assert len(json_paths) == len(selected)

    schema = _load_schema()
    for json_path in json_paths:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        errors = _validate_schema(payload, schema)
        assert not errors, f"{json_path.name} schema errors: {errors}"

    validation_dir = temp_out / "validation"
    validation_cmd = [
        sys.executable,
        "-m",
        "metricas_lattes.exports.validation_pack",
        "--in",
        str(researchers_dir),
        "--out",
        str(validation_dir),
        "--format",
        "html",
        "xlsx",
    ]
    validation_run = subprocess.run(validation_cmd, capture_output=True, text=True)
    assert validation_run.returncode == 0, validation_run.stdout + validation_run.stderr

    html_files = list(validation_dir.glob("researchers/*/VALIDACAO.html"))
    xlsx_files = list(validation_dir.glob("researchers/*/VALIDACAO.xlsx"))
    assert len(html_files) == len(selected)
    assert len(xlsx_files) == len(selected)

    author_pattern = re.compile(r"^[A-ZÀ-Ü]{2,},\s*[A-ZÀ-Ü]")
    for html_path in html_files:
        html_text = html_path.read_text(encoding="utf-8")
        rows = re.findall(r"<tr>(.*?)</tr>", html_text, flags=re.DOTALL)
        inspected = 0
        for row in rows:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row, flags=re.DOTALL)
            if len(cells) < 3:
                continue
            titulo = html_lib.unescape(re.sub(r"<[^>]+>", "", cells[1])).strip()
            if not titulo:
                continue
            inspected += 1
            assert " ; " not in titulo, f"Titulo com separador em HTML: {titulo}"
        assert inspected > 0, f"Nenhum titulo encontrado em {html_path}"

    total_target_rows = 0
    for xlsx_path in xlsx_files:
        from openpyxl import load_workbook

        workbook = load_workbook(xlsx_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        col_index = {name: idx for idx, name in enumerate(headers)}

        titulo_col = col_index.get("titulo")
        autores_col = col_index.get("autores")
        section_col = col_index.get("section")

        assert titulo_col is not None
        assert autores_col is not None
        assert section_col is not None

        target_rows = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            titulo = row[titulo_col]
            autores = row[autores_col]
            section = row[section_col]

            if titulo:
                assert " ; " not in str(titulo), f"Titulo com separador em XLSX: {titulo}"

            normalized_section = _normalize_section(str(section or ""))
            if normalized_section in TARGET_SECTIONS:
                target_rows += 1
                if titulo:
                    assert not author_pattern.search(str(titulo)), f"Titulo com autores em XLSX: {titulo}"
                assert autores, f"Autores vazio em XLSX para secao {section}: {titulo}"

        total_target_rows += target_rows

    assert total_target_rows > 0, "Nenhuma linha das secoes alvo nos XLSXs"
