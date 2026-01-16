"""Tests for validation pack XLSX exports."""

from pathlib import Path
import json
import re
import shutil

from openpyxl import load_workbook

from metricas_lattes.exports.validation_pack import COLUMN_ORDER, generate_validation_pack


FIXTURE_PATH = Path('tests/fixtures/validation_pack/carlos_alberto_perez.json')


def _is_year(value) -> bool:
    if value is None:
        return False
    if isinstance(value, int):
        return 1900 <= value <= 2100
    if isinstance(value, str):
        return re.match(r'^(19|20)\d{2}$', value.strip()) is not None
    return False


def test_xlsx_column_alignment(tmp_path: Path) -> None:
    input_dir = tmp_path / 'input'
    output_dir = tmp_path / 'out'
    input_dir.mkdir()

    fixture_copy = input_dir / FIXTURE_PATH.name
    shutil.copyfile(FIXTURE_PATH, fixture_copy)

    generate_validation_pack(input_dir, output_dir, ['xlsx'])

    data = json.loads(fixture_copy.read_text(encoding='utf-8'))
    lattes_id = data['researcher']['lattes_id']
    xlsx_path = output_dir / 'researchers' / f'{lattes_id}__' / 'VALIDACAO.xlsx'
    assert xlsx_path.exists()

    workbook = load_workbook(xlsx_path, data_only=True)
    assert 'Produções' in workbook.sheetnames

    sheet = workbook['Produções']
    header = [cell.value for cell in sheet[1]]
    assert header == COLUMN_ORDER

    first_row = [cell.value for cell in sheet[2]]
    col_index = {name: idx for idx, name in enumerate(header)}

    ano_value = first_row[col_index['ano']]
    titulo_value = first_row[col_index['titulo']]
    autores_value = first_row[col_index['autores']]

    assert _is_year(ano_value)
    assert isinstance(titulo_value, str) and titulo_value.strip()
    assert titulo_value != autores_value


def _find_row_by_title(sheet, header, needle: str):
    col_index = {name: idx for idx, name in enumerate(header)}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        titulo = row[col_index['titulo']]
        if isinstance(titulo, str) and needle in titulo:
            return row, col_index
    return None, col_index


def test_validation_pack_field_fallbacks(tmp_path: Path) -> None:
    input_dir = tmp_path / 'input'
    output_dir = tmp_path / 'out'
    input_dir.mkdir()

    fixture_copy = input_dir / FIXTURE_PATH.name
    shutil.copyfile(FIXTURE_PATH, fixture_copy)

    generate_validation_pack(input_dir, output_dir, ['xlsx'])

    data = json.loads(fixture_copy.read_text(encoding='utf-8'))
    lattes_id = data['researcher']['lattes_id']
    xlsx_path = output_dir / 'researchers' / f'{lattes_id}__' / 'VALIDACAO.xlsx'

    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook['Produções']
    header = [cell.value for cell in sheet[1]]

    row, col_index = _find_row_by_title(sheet, header, 'SiC Structural Analysis')
    assert row is not None
    autores_value = row[col_index['autores']]
    veiculo_value = row[col_index['veiculo_ou_livro']]
    assert isinstance(autores_value, str)
    assert re.search(r'\b(19|20)\d{2}\b', autores_value) is None
    assert isinstance(veiculo_value, str) and veiculo_value.strip()

    row, col_index = _find_row_by_title(sheet, header, 'THE TXRF TECHNIQUE AT THE LNLS')
    assert row is not None
    titulo_value = row[col_index['titulo']]
    assert isinstance(titulo_value, str) and titulo_value.strip()
